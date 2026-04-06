"""
match_invoices.py
-----------------
Reads the Market Rate Key and Invoices report from a Statewide Interpreters
invoice workbook (.xlsx) and classifies each invoice by market rate type.

Outputs a CSV file alongside the workbook: <workbook_name>_matched.csv
Columns: InvoiceNum, RateType, Memo

Usage:
    python match_invoices.py "path\\to\\invoices workbook.xlsx"

Notes:
    - Does NOT modify the Excel file.
    - RateType values: Type 1–9, '?' (unmatched service), or blank (supplementary/excluded).
    - '?' rows are printed at the end so you can add them to the Market Rate Key.
"""

import sys
import re
import csv
import os
import struct
import shutil
import tempfile
import zipfile
from openpyxl import load_workbook

# ── Column indices (1-based Excel columns → 0-based Python) ──────────────────
MEMO_COL =  4 - 1   # Column D
ADDR_COL =  3 - 1   # Column C (Name Address)
NUM_COL  =  5 - 1   # Column E (Invoice number)
RATE_COL =  8       # Column H (1-based, for openpyxl write)


# ── Supplementary row detection ───────────────────────────────────────────────
SUPPLEMENTARY_PATTERNS = [
    re.compile(r'^99031',              re.IGNORECASE),
    re.compile(r'^parking',            re.IGNORECASE),
    re.compile(r'^INV\s*#',           re.IGNORECASE),
    re.compile(r'^Agency fees',        re.IGNORECASE),
    re.compile(r'^Travel Time',        re.IGNORECASE),
    re.compile(r'^VOID:',              re.IGNORECASE),
    re.compile(r'^Petition For Cost',  re.IGNORECASE),
]
SUPPLEMENTARY_EXACT = {'IBR Process'}

CERT_PATTERN = re.compile(
    r'Certified Medical Interpreter|CCHI|CMI\s*#|#\s*CMI|NBCMI|'
    r'Medical Certif|Interpreter\s+Medical\s+Cert|'
    r'Administrative.*Certified|Administrative.*Interpreter|'
    r'Provisionally|provisionally|certified interpreter|'
    r'Court Certified Interpreter|Certification\s*#|'
    r'#\s*CHI\s*\d|#\s*\d{4,}',
    re.IGNORECASE
)

HOURS_PATTERN = re.compile(r'^\d[\d\s/\.]*\s+Hours?', re.IGNORECASE)


def is_supplementary(memo: str) -> bool:
    if memo in SUPPLEMENTARY_EXACT:
        return True
    for pat in SUPPLEMENTARY_PATTERNS:
        if pat.match(memo):
            return True
    if CERT_PATTERN.search(memo):
        return True
    if ':' in memo and re.search(r'Interpreter\.?\s*$', memo, re.IGNORECASE):
        return True
    return False


def load_market_rate_key(ws_key):
    """
    Returns:
        type_data  : dict  { type_label: {'filter': str|None, 'memos': [str, ...]} }
                     'filter' is the Col E Address Filter value (or None if blank).
                     A type with a filter only matches invoices whose Name Address
                     contains that filter string — see classify_memo().
        exclusions : list  [ name_address_string, ... ]

    Market Rate Key column layout:
        Col A  Type label        e.g. "Type 8"
        Col B  Description
        Col C  Rate
        Col D  MR PDF label
        Col E  Address Filter    e.g. "alliedmanagedcare.com"  (blank = no filter)
        Col F+ Memo descriptions (one per cell, same row)

    Handles the edge case where Col A is blank but Col D contains the type label
    (e.g. a row where Col A was accidentally left empty).
    """
    type_data  = {}
    exclusions = []

    for row in ws_key.iter_rows(min_row=2):
        a_val = row[0].value
        d_val = row[3].value if len(row) > 3 else None

        # Determine label: prefer Col A; fall back to Col D if it matches "Type N"
        if a_val:
            label = str(a_val).strip()
        elif d_val and re.match(r'^Type\s*\d+$', str(d_val).strip(), re.IGNORECASE):
            label = str(d_val).strip()
        else:
            continue

        if label.lower().startswith('list of'):
            for cell in row[1:]:
                if cell.value:
                    exclusions.append(str(cell.value).strip())
            continue

        # Col E (index 4): optional address filter
        e_val = row[4].value if len(row) > 4 else None
        addr_filter = str(e_val).strip() if e_val else None

        # Col F onward (index 5+): memo descriptions
        memos = [str(cell.value).strip() for cell in row[5:] if cell.value]

        if memos:
            type_data[label] = {'filter': addr_filter, 'memos': memos}

    return type_data, exclusions


def classify_memo(memo: str, addr: str, type_data: dict) -> str | None:
    """
    Returns type label, '?', or None (blank / supplementary).
    Priority:
      1. Supplementary  → None
      2. Key match — filtered types first (exact, then 60-char prefix)
      3. Key match — unfiltered types (exact, then 60-char prefix)
      4. Pattern fallbacks (ordered by specificity)
      5. Anything else that looks like a service row → '?'

    type_data : { label: {'filter': str|None, 'memos': [str]} }
        Types with a non-None 'filter' only match when that string appears
        in the invoice's Name Address (Column L).  This is how Type 8
        (Allied Cantonese-Mandarin) is distinguished from Type 7 —
        the Market Rate Key carries 'alliedmanagedcare.com' in Col E.

    addr : Name Address column value (Column L).
    """
    m = memo.strip()
    if not m:
        return None

    if is_supplementary(m):
        return None

    # ── Key-based matching ────────────────────────────────────────────────────
    # Pass 1: types WITH an address filter (more specific — checked first)
    for label, entry in type_data.items():
        if not entry['filter']:
            continue
        if entry['filter'].lower() not in addr.lower():
            continue
        for key_memo in entry['memos']:
            if m == key_memo:
                return label
    for label, entry in type_data.items():
        if not entry['filter']:
            continue
        if entry['filter'].lower() not in addr.lower():
            continue
        for key_memo in entry['memos']:
            if m[:60] == key_memo[:60]:
                return label

    # Pass 2: types WITHOUT an address filter
    for label, entry in type_data.items():
        if entry['filter']:
            continue
        for key_memo in entry['memos']:
            if m == key_memo:
                return label
    for label, entry in type_data.items():
        if entry['filter']:
            continue
        for key_memo in entry['memos']:
            if m[:60] == key_memo[:60]:
                return label

    # ── Pattern fallbacks (ordered by specificity) ────────────────────────────

    # Type 3: C&R and Stipulation readings
    if re.search(r'Reading of a Compromise and Release|Stipulation & Award reading', m, re.IGNORECASE):
        return 'Type 3'

    # Type 4: Deposition transcript readings
    if re.search(r'Reading of a deposition transcript', m, re.IGNORECASE):
        return 'Type 4'

    # Half-day / Full-day — match language to legal type
    if re.search(r'Half\s+(a\s+)?Day|Full\s+Day', m, re.IGNORECASE):
        if re.search(r'Cantonese|Mandarin', m, re.IGNORECASE):
            return 'Type 6'   # Cantonese/Mandarin legal
        if re.search(r'Spanish', m, re.IGNORECASE):
            return 'Type 5'   # Spanish legal
        if re.search(r'Tagalog', m, re.IGNORECASE):
            return 'Type 14'  # Tagalog legal
        if re.search(r'Arabic', m, re.IGNORECASE):
            return 'Type 15'  # Arabic legal
        return '?'   # Other language half/full day — needs a new type

    # Type 2: Panel QME / AME (Spanish)
    if re.search(r'Panel\s+QME|\bAME\b', m, re.IGNORECASE):
        return 'Type 2'

    # Hours-based language matching
    if HOURS_PATTERN.match(m) and 'interpreting services' in m.lower():
        lang = m.lower()
        is_allied = 'alliedmanagedcare.com' in addr.lower()

        if 'vietnamese' in lang:
            return 'Type 9'
        if 'cantonese' in lang or 'mandarin' in lang:
            return 'Type 8' if is_allied else 'Type 7'  # Allied vs. standard Cantonese/Mandarin
        if 'tagalog' in lang:
            return 'Type 11'
        if 'korean' in lang:
            return 'Type 12'
        if 'arabic' in lang:
            return 'Type 13'
        # Type 10: all other non-Spanish languages
        OTHER_LANGUAGES = ['farsi', 'laotian', 'punjabi', 'armenian', 'fijian',
                           'hindi', 'egyptian', 'russian', 'urdu', 'asl', 'tongan']
        if any(x in lang for x in OTHER_LANGUAGES):
            return 'Type 10'
        if 'spanish' in lang or (
            'certified' in lang and
            not any(x in lang for x in ['cantonese', 'mandarin', 'chinese', 'korean',
                                         'vietnamese', 'farsi', 'arabic', 'tagalog',
                                         'hindi', 'punjabi', 'russian', 'armenian',
                                         'laotian', 'fijian', 'tongan', 'urdu', 'asl'])
        ):
            return 'Type 1'

        return '?'   # Unrecognized language — add to Market Rate Key

    if HOURS_PATTERN.match(m):
        return '?'

    # Remaining pending patterns
    pending_patterns = [
        r'for\s+(a\s+|an\s+)?(Meeting|Telephone\s+meeting|Employee\s+Meeting|Settlement\s+Agreement|Discovery\s+Questions)',
        r'translation\s+documents',
    ]
    for pat in pending_patterns:
        if re.search(pat, m, re.IGNORECASE):
            return '?'

    return None


def repair_xlsx_if_needed(path: str) -> str:
    """
    If the xlsx has a missing end-of-central-directory (e.g. closed mid-save or
    OneDrive still syncing), rebuild it from the central directory entries that are
    present and return a path to a repaired temp copy. Otherwise returns path unchanged.
    """
    with open(path, 'rb') as f:
        data = f.read()

    # Fast check: does a valid EOCD already exist?
    if data.rfind(b'PK\x05\x06') != -1:
        return path

    print("  Warning: workbook missing end-of-central-directory — attempting repair.")
    print("  (Close the file in Excel before running this script.)")

    first_cd = data.find(b'PK\x01\x02')
    if first_cd == -1:
        raise ValueError("Cannot repair: no central directory entries found. Ensure Excel is closed.")

    pos, valid_end = first_cd, first_cd
    num_entries = 0
    while pos < len(data) - 46 and data[pos:pos+4] == b'PK\x01\x02':
        try:
            fields = struct.unpack('<4s6H3I5H2I', data[pos:pos+46])
            fname_len, extra_len, comment_len = fields[10], fields[11], fields[12]
            pos += 46 + fname_len + extra_len + comment_len
            valid_end = pos
            num_entries += 1
        except Exception:
            break

    cd_start = first_cd
    cd_size  = valid_end - cd_start
    eocd = struct.pack('<4sHHHHIIH', b'PK\x05\x06', 0, 0,
                       num_entries, num_entries, cd_size, cd_start, 0)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.write(data[:valid_end])
    tmp.write(eocd)
    tmp.close()

    # Verify
    try:
        zipfile.ZipFile(tmp.name).close()
    except Exception as e:
        os.unlink(tmp.name)
        raise ValueError(f"Repair failed: {e}. Please ensure Excel is closed and OneDrive has finished syncing.")

    print(f"  Repair successful ({num_entries} entries recovered).")
    return tmp.name


def process_workbook(path: str):
    print(f"Reading: {path}")
    repaired_path = repair_xlsx_if_needed(path)

    # Open writable so we can populate Column H (MR Type)
    tmp = path + '.tmp.xlsx'
    shutil.copy2(repaired_path, tmp)
    if repaired_path != path:
        os.unlink(repaired_path)

    wb = load_workbook(tmp)
    ws_inv = wb['Invoice Key']
    ws_key = wb['Market Rate Key']

    type_data, exclusions = load_market_rate_key(ws_key)
    print(f"  Loaded {len(type_data)} market rate types: {', '.join(sorted(type_data))}")
    filtered = [l for l, e in type_data.items() if e['filter']]
    if filtered:
        print(f"  Address-filtered types: {', '.join(filtered)}")
    print(f"  Loaded {len(exclusions)} exclusion address(es).")

    invoice_results = {}   # { invoice_num: (rate_type, memo, addr) }
    counts = {}
    question_memos = []

    for row in ws_inv.iter_rows(min_row=2):
        num_cell  = row[NUM_COL]
        memo_cell = row[MEMO_COL]
        addr_cell = row[ADDR_COL]
        rate_cell = row[RATE_COL - 1]  # Convert 1-based to 0-based for row tuple

        invoice_num = str(num_cell.value).strip() if num_cell.value else ''
        if not invoice_num:
            continue

        memo = str(memo_cell.value).strip() if memo_cell.value else ''
        addr = str(addr_cell.value).strip() if addr_cell.value else ''

        # Exclusion check
        if any(excl and (excl in addr or addr in excl) for excl in exclusions):
            rate_cell.value = 'excluded'
            if invoice_num not in invoice_results:
                invoice_results[invoice_num] = ('excluded', memo, addr)
            counts['excluded'] = counts.get('excluded', 0) + 1
            continue

        if not memo:
            continue

        result = classify_memo(memo, addr, type_data)

        # Write result to Column H
        rate_cell.value = result

        if result is not None:
            if invoice_num not in invoice_results or invoice_results[invoice_num][0] in (None, ''):
                invoice_results[invoice_num] = (result, memo, addr)
            key = result
        else:
            if invoice_num not in invoice_results:
                invoice_results[invoice_num] = (None, memo, addr)
            key = 'blank'

        counts[key] = counts.get(key, 0) + 1
        if result == '?':
            question_memos.append(memo[:80])

    # Save workbook with Column H populated
    wb.save(tmp)
    os.replace(tmp, path)
    print(f"\n  Column H (MR Type) written to workbook.")

    # Also write CSV (same directory as workbook)
    base = os.path.splitext(os.path.basename(path))[0]
    csv_path = os.path.join(os.path.dirname(path), base + '_matched.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['InvoiceNum', 'RateType', 'Memo', 'NameAddress'])
        for inv_num in sorted(invoice_results.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            rate, memo, addr = invoice_results[inv_num]
            writer.writerow([inv_num, rate if rate else '', memo, addr])

    print(f"  Output CSV: {csv_path}")
    print(f"  Total unique invoices: {len(invoice_results)}")

    print("\nClassification summary:")
    for k, v in sorted(counts.items()):
        print(f"  {k:12s}: {v}")

    if question_memos:
        unique_q = sorted(set(question_memos))
        print(f"\n'?' memos needing attention ({len(unique_q)} unique):")
        for q in unique_q:
            print(f"  {q}")

    wb.close()
    return csv_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python match_invoices.py <workbook.xlsx>")
        sys.exit(1)
    process_workbook(sys.argv[1])
