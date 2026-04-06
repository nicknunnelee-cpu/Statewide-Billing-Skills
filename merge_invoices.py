"""
merge_invoices.py
-----------------
Reads the CLASSIFIED invoice workbook, then for every invoice PDF in the
source folder:

  1. Copies it to the source folder renamed as "{NameAddress} {InvoiceNum}.pdf"
     (all invoices, whether or not they have a Rate Type).
  2. If Column T contains a Rate Type (Type 1–15):
       merges the renamed invoice PDF with the matching "MR type N.pdf" and
       saves the merged file to the output folder with the same name.

Usage (run from Command Prompt):
    python merge_invoices.py ^
        "C:\\...\\3.25.26 invoices to send, all of March xl - CLASSIFIED.xlsx" ^
        "C:\\...\\Inv PDFs 3.26.26" ^
        "C:\\...\\Nico Data" ^
        "C:\\...\\Merged 3.26.26"

Arguments:
    classified_xlsx   Path to the CLASSIFIED workbook (Invoices report + Col T)
    inv_pdfs_folder   Folder containing raw invoice PDFs named by invoice number
    mr_pdfs_folder    Folder containing "MR type 1.pdf" … "MR type 15.pdf"
    output_folder     Where merged PDFs are saved
"""

import sys
import os
import re
import shutil
import struct
import tempfile
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

# ── Column indices (0-based) ──────────────────────────────────────────────────
NUM_COL  =  5 - 1   # Col E  Invoice number
ADDR_COL =  3 - 1   # Col C  Name Address
RATE_COL =  8 - 1   # Col H  Rate Type

# Characters not allowed in Windows filenames
_INVALID_CHARS = re.compile(r'[\\/*?:"<>|]')


def sanitize_filename(text: str) -> str:
    """Strip surrounding whitespace/dots and remove Windows-illegal characters."""
    text = text.strip().rstrip('.')
    return _INVALID_CHARS.sub('', text)


def repair_xlsx_if_needed(path: str) -> str:
    """
    Repairs an xlsx that is missing its end-of-central-directory record or
    entire central directory (common when OneDrive is mid-sync or Excel had
    the file open).  Returns the original path if already valid, otherwise
    writes a repaired copy to a temp file and returns that path.

    Strategy A – central directory exists but EOCD missing:
        Rebuild EOCD from the existing PK\x01\x02 entries.
    Strategy B – no central directory at all:
        Scan for PK\x03\x04 local file headers, decompress each entry with
        zlib, and reassemble as a fresh valid ZIP/xlsx.
    """
    import zlib

    with open(path, 'rb') as f:
        data = f.read()

    if data.rfind(b'PK\x05\x06') != -1:
        return path  # already valid

    print("  Warning: workbook missing end-of-central-directory — attempting repair.")
    print("  (Make sure Excel is closed before running this script.)")

    # ── Strategy A: rebuild EOCD from existing central-directory entries ──────
    first_cd = data.find(b'PK\x01\x02')
    if first_cd != -1:
        pos, valid_end, n = first_cd, first_cd, 0
        while pos < len(data) - 46 and data[pos:pos+4] == b'PK\x01\x02':
            try:
                fields = struct.unpack('<4s6H3I5H2I', data[pos:pos+46])
                pos += 46 + fields[10] + fields[11] + fields[12]
                valid_end = pos
                n += 1
            except Exception:
                break

        eocd = struct.pack('<4sHHHHIIH', b'PK\x05\x06', 0, 0,
                           n, n, valid_end - first_cd, first_cd, 0)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.write(data[:valid_end])
        tmp.write(eocd)
        tmp.close()

        try:
            zipfile.ZipFile(tmp.name).close()
            print(f"  Repair successful (Strategy A, {n} entries recovered).")
            return tmp.name
        except Exception:
            os.unlink(tmp.name)
            # fall through to Strategy B

    # ── Strategy B: reconstruct from local file headers ──────────────────────
    print("  No central directory found — rebuilding from local headers.")
    entries = []
    pos = 0
    while pos < len(data) - 30:
        idx = data.find(b'PK\x03\x04', pos)
        if idx == -1:
            break
        pos = idx
        if pos + 30 > len(data):
            break
        try:
            ver, flag, method, mod_time, mod_date, crc, comp_sz, uncomp_sz, fname_len, extra_len = \
                struct.unpack_from('<5H3I2H', data, pos + 4)
        except struct.error:
            pos += 4
            continue

        hdr_end = pos + 30 + fname_len + extra_len
        fname = data[pos + 30: pos + 30 + fname_len].decode('utf-8', errors='replace')

        if hdr_end + comp_sz > len(data):
            pos += 4
            continue

        raw = data[hdr_end: hdr_end + comp_sz]

        if method == 0:
            content = raw
        elif method == 8:
            try:
                content = zlib.decompress(raw, -15)
            except zlib.error:
                pos = hdr_end + comp_sz
                continue
        else:
            pos = hdr_end + comp_sz
            continue

        entries.append((fname, content))
        pos = hdr_end + comp_sz

    if not entries:
        raise ValueError(
            "Cannot repair: no recoverable entries found. "
            "Ensure Excel is closed and OneDrive has finished syncing."
        )

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    with zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zout:
        for fname, content in entries:
            zout.writestr(fname, content)

    try:
        zipfile.ZipFile(tmp.name).close()
        print(f"  Repair successful (Strategy B, {len(entries)} entries recovered).")
        return tmp.name
    except Exception as e:
        os.unlink(tmp.name)
        raise ValueError(f"Repair failed: {e}")


def build_invoice_map(xlsx_path: str) -> dict:
    """
    Returns { invoice_num_str: {'addr': str, 'rate': str} }
    Reads Invoices report sheet; picks the first non-blank addr and
    first non-blank rate found for each invoice number.
    """
    repaired = repair_xlsx_if_needed(xlsx_path)
    wb = load_workbook(repaired, data_only=True, read_only=True)
    if repaired != xlsx_path:
        os.unlink(repaired)

    ws = wb['Invoice Key']
    invoice_map = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        num  = str(row[NUM_COL]).strip()  if row[NUM_COL]  else ''
        addr = str(row[ADDR_COL]).strip() if row[ADDR_COL] else ''
        rate = str(row[RATE_COL]).strip() if row[RATE_COL] else ''
        if not num or num == 'None':
            continue
        if num not in invoice_map:
            invoice_map[num] = {'addr': addr, 'rate': rate}
        else:
            if not invoice_map[num]['addr'] and addr:
                invoice_map[num]['addr'] = addr
            if not invoice_map[num]['rate'] and rate:
                invoice_map[num]['rate'] = rate

    wb.close()
    return invoice_map


def get_mr_type_number(rate_type: str):
    """Extract integer from 'Type 7' → 7, or return None."""
    m = re.match(r'^Type\s*(\d+)$', rate_type.strip(), re.IGNORECASE)
    return int(m.group(1)) if m else None


def merge_pdfs(invoice_pdf: str, mr_pdf: str, output_path: str):
    writer = PdfWriter()
    for path in [invoice_pdf, mr_pdf]:
        for page in PdfReader(path).pages:
            writer.add_page(page)
    with open(output_path, 'wb') as f:
        writer.write(f)


def main():
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    xlsx_path      = sys.argv[1]
    inv_pdfs_dir   = sys.argv[2]
    mr_pdfs_dir    = sys.argv[3]
    output_dir     = sys.argv[4]

    for p in [xlsx_path, inv_pdfs_dir, mr_pdfs_dir]:
        if not os.path.exists(p):
            print(f"ERROR: path not found: {p}")
            sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    # ── Build invoice map from Excel ──────────────────────────────────────────
    print(f"Reading: {xlsx_path}")
    invoice_map = build_invoice_map(xlsx_path)
    print(f"  {len(invoice_map)} invoices loaded from workbook.")

    # ── Process each invoice PDF ──────────────────────────────────────────────
    pdfs = sorted(f for f in os.listdir(inv_pdfs_dir) if f.lower().endswith('.pdf'))
    print(f"\nProcessing {len(pdfs)} invoice PDFs from: {inv_pdfs_dir}\n")

    copied   = 0
    merged   = 0
    skipped  = 0
    warnings = []

    for fname in pdfs:
        # Invoice number = first token of the filename stem
        stem = os.path.splitext(fname)[0]
        inv_num = stem.split()[0]

        entry = invoice_map.get(inv_num)
        if not entry:
            warnings.append(f"No Excel entry for PDF: {fname}")
            skipped += 1
            continue

        addr      = sanitize_filename(entry['addr']) if entry['addr'] else ''
        rate_type = entry['rate']

        # Build new filename — address first so files sort alphabetically by client
        new_stem = f"{addr} {inv_num}".strip() if addr else inv_num
        new_name = f"{new_stem}.pdf"

        src_path    = os.path.join(inv_pdfs_dir, fname)
        output_path = os.path.join(output_dir, new_name)

        # ── Step 1: Copy renamed invoice into the output folder (all invoices) ─
        if os.path.exists(output_path):
            print(f"  Already in output: {new_name}")
        else:
            shutil.copy2(src_path, output_path)
            copied += 1
            print(f"  Copied  : {fname}")
            print(f"        → {new_name}")

        # ── Step 2: If rate type exists, merge the output copy with MR PDF ────
        type_num = get_mr_type_number(rate_type) if rate_type else None
        if type_num is None:
            continue  # no rate type — renamed copy only, no merge

        mr_filename = f"MR type {type_num}.pdf"
        mr_path     = os.path.join(mr_pdfs_dir, mr_filename)
        if not os.path.exists(mr_path):
            warnings.append(f"MR PDF not found: {mr_filename} (invoice #{inv_num})")
            skipped += 1
            continue

        try:
            # Merge: overwrite the output copy with invoice + MR pages combined
            merged_tmp = output_path + ".tmp.pdf"
            merge_pdfs(output_path, mr_path, merged_tmp)
            os.replace(merged_tmp, output_path)
            print(f"  Merged  : {new_name}  +  {mr_filename}")
            merged += 1
        except Exception as e:
            warnings.append(f"Merge error for #{inv_num}: {e}")
            skipped += 1

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'─'*60}")
    print(f"  Copied  : {copied}")
    print(f"  Merged  : {merged}")
    print(f"  Skipped : {skipped}")
    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings:
            print(f"  ✗ {w}")


if __name__ == '__main__':
    main()
